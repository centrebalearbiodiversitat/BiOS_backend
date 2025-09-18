import traceback

from openpyxl import load_workbook
from django.core.management.base import BaseCommand
from django.db import transaction
from apps.taxonomy.models import TaxonomicLevel
from apps.tags.models import Tag, TaxonTag, System, Directive
from apps.versioning.models import Batch, OriginId, Source
from common.utils.utils import get_or_create_source, is_batch_referenced
from tqdm import tqdm

EXTERNAL_ID = "origin_id"
INTERNAL_NAME = "source"
SOURCE_TYPE = "origin"
SOURCE_METHOD = "extraction_method"


def check_taxon(line):
	taxonomy = TaxonomicLevel.objects.find(
		taxon=line["origin_taxon"]
	).filter(rank=TaxonomicLevel.TRANSLATE_RANK[line["taxon_rank"]])

	if taxonomy.count() == 0:
		raise Exception(f"Taxonomy not found.")
	elif taxonomy.count() > 1:
		raise Exception(f"Multiple taxonomy found.")

	return taxonomy.first()


def parse_bool(value, return_none=False):
	if value is None:
		if return_none:
			return None
		else:
			raise Exception("Bool value cannot be None")
	else:
		if isinstance(value, bool):
			return value
		elif value.lower() == "true":
			return True
		elif value.lower() == "false":
			return False
		else:
			raise Exception(f"Invalid boolean value: {value}")


def load_taxon_tags(line, taxonomy, batch):
	if line["taxon_rank"] not in ["species", "subspecies", "variety"]:
		raise Exception(f"Taxon rank not allowed.")

	source = get_or_create_source(
		source_type=line[SOURCE_TYPE].lower(),
		extraction_method=Source.EXPERT,
		data_type=Source.TAXON_DATA,
		batch=batch,
		internal_name=line[INTERNAL_NAME],
	)

	# try:
	system, _ = System.objects.get_or_create(taxonomy=taxonomy, defaults={"batch": batch})
	# except System.DoesNotExist:
	# 	raise Exception(f"Priority check: Taxon systems not found. Taxon data (iucn) not loaded yet? Otherwise iucn will overwrite expert data")

	os, _ = OriginId.objects.get_or_create(source=source, defaults={"attribution": line.get("attribution")})
	# if the 3 systems are None, it means IUCN has no available data
	if system.freshwater is None and system.marine is None and system.terrestrial is None:
		system.freshwater = parse_bool(line["freshwater"])
		system.marine = parse_bool(line["marine"])
		system.terrestrial = parse_bool(line["terrestrial"])
		system.sources.clear()
		system.sources.add(os)
		system.save()

	doe_value = line.get("degreeOfEstablishment")
	try:
		doe_tag = Tag.objects.get(name__iexact=doe_value, tag_type=Tag.DOE)
		taxon_tag, _ = TaxonTag.objects.get_or_create(
			taxonomy=taxonomy,
			defaults={
				"batch": batch,
				"tag": doe_tag,
			},
		)
		taxon_tag.sources.add(os)
	except Tag.DoesNotExist:
		raise Exception(f"No Tag.DOE was found with the value '{doe_value}'")

	directive, _ = Directive.objects.get_or_create(
		taxonomy=taxonomy,
		defaults={
			"cites": parse_bool(line["cites"], True),
			"ceea": parse_bool(line["ceea"], True),
			"lespre": parse_bool(line["lespre"], True),
			"directiva_aves": parse_bool(line["directiva_aves"], True),
			"directiva_habitats": parse_bool(line["directiva_habitats"], True),
			"batch": batch,
		},
	)
	directive.sources.add(os)


class Command(BaseCommand):
	def add_arguments(self, parser):
		parser.add_argument("file", type=str, help="Path to the data file")
		parser.add_argument("-d", nargs="?", type=str, default=";")

	@transaction.atomic
	def handle(self, *args, **options):
		file_name = options["file"]
		exception = False
		batch = Batch.objects.create()

		try:
			reader = load_workbook(file_name)
			sheet = reader.active
			headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
			for row in tqdm(list(sheet.iter_rows(min_row=2, values_only=True)), ncols=50, colour="yellow", smoothing=0, miniters=100, delay=20):
				line = dict(zip(headers, row))
				if line["origin_taxon"] is None:
					continue
				try:
					taxonomy = check_taxon(line)
					load_taxon_tags(line, taxonomy, batch)
				except:
					exception = True
					print(traceback.format_exc(), line)
		except FileNotFoundError:
			exception = True
			print("No such file or directory")

		if exception:
			raise Exception(f"Errors found: Rollback control")

		is_batch_referenced(batch)
